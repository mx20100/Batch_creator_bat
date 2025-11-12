import os
import sys
import csv
import zipfile
import shutil
import logging
import io
from openpyxl import load_workbook
import tempfile
from datetime import datetime

# ------------------------------------------------------------
# Cooperative cancellation
# ------------------------------------------------------------
CANCEL_FLAG = False

def request_cancel():
    global CANCEL_FLAG
    CANCEL_FLAG = True

def check_cancel():
    if CANCEL_FLAG:
        raise RuntimeError("Conversion cancelled by user.")

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------
REQUIRED_COLUMNS = [
    "batch", "filename", "material", "part_id",
    "copies", "next_step", "order_id", "technology"
]

ENCODING = "utf-8"
MAX_ZIP_SIZE_MB = 900
MAX_ZIP_SIZE_BYTES = MAX_ZIP_SIZE_MB * 1024 * 1024

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def get_working_dir() -> str:
    return os.getcwd()

def setup_logger(log_path: str | None = None, external_logger: logging.Logger | None = None) -> logging.Logger:
    """
    Initialize or reuse a logger for the converter.
    - If an external logger is provided (from GUI), reuse it as-is.
    - If no external logger, create a file-based one.
    """
    if external_logger:
        logger = external_logger
        # Only log startup banner if this is a fresh session
        if not any(isinstance(h, logging.FileHandler) for h in logger.handlers):
            logger.info("=" * 60)
            logger.info("Converter started (attached to external logger)")
        return logger

    # Create or reuse a named logger
    logger = logging.getLogger("converter")
    if logger.handlers:
        return logger  # already configured by GUI

    logger.setLevel(logging.INFO)

    # Only add file handler if a path is provided
    if log_path:
        fh = logging.FileHandler(log_path, encoding=ENCODING)
        fh.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        fh.setFormatter(formatter)
        logger.addHandler(fh)

    logger.info("=" * 60)
    logger.info("Converter started (standalone mode)")
    return logger

# ------------------------------------------------------------
# Excel → CSV (in-memory)
# ------------------------------------------------------------
def convert_excel_to_csv_buffer(excel_path: str, logger: logging.Logger) -> io.StringIO:
    """Reads Excel and returns its meta.csv as an in-memory buffer."""
    import pandas as pd
    logger.info(f"Converting Excel to CSV (in-memory): {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    removed = 0
    for sheet in wb.sheetnames[:]:
        ws = wb[sheet]
        if not any(any(cell not in (None, "", " ") for cell in row) for row in ws.iter_rows(values_only=True)):
            wb.remove(ws)
            logger.info(f"Removed empty sheet: {sheet}")
            removed += 1

    import tempfile

    if removed:
        # Save cleaned workbook into a hidden temporary file in the system temp directory
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            temp_cleaned_path = tmp_file.name

        try:
            wb.save(temp_cleaned_path)
            wb.close()
            # Reopen the cleaned workbook safely
            wb = load_workbook(temp_cleaned_path, data_only=True, read_only=True)

        finally:
            # Delete the cleaned file immediately after (user never sees it)
            try:
                os.remove(temp_cleaned_path)
            except OSError:
                pass

    ws = next((s for s in wb.worksheets if s.sheet_state == "visible"), None)
    if ws is None:
        logger.warning("No visible sheets — using pandas fallback.")
        df = pd.read_excel(excel_path, engine="openpyxl")
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding=ENCODING)
        csv_buffer.seek(0)
        return csv_buffer

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    csv_buffer.seek(0)
    logger.info("Conversion to meta.csv completed (in-memory).")
    return csv_buffer

# ------------------------------------------------------------
# CSV validation
# ------------------------------------------------------------
def validate_and_fix_meta_buffer(csv_buffer: io.StringIO, logger: logging.Logger) -> io.StringIO:
    """Validates and fixes meta.csv content, returning updated in-memory buffer."""
    csv_buffer.seek(0)
    reader = csv.DictReader(csv_buffer)
    rows = list(reader)
    header = reader.fieldnames
    if header is None:
        raise ValueError("meta.csv missing header row")

    if [h.strip().lower() for h in header[: len(REQUIRED_COLUMNS)]] != REQUIRED_COLUMNS:
        raise ValueError(f"Header mismatch. Found: {header}, Expected: {REQUIRED_COLUMNS}")

    fixed_copies = fixed_filenames = 0
    for row in rows:
        check_cancel()
        if (row.get("copies") or "").strip() in ("", "0"):
            row["copies"] = "1"
            fixed_copies += 1
        fname = (row.get("filename") or "").strip()
        if fname and not fname.lower().endswith(".stl"):
            row["filename"] = fname + ".stl"
            fixed_filenames += 1

    out_buf = io.StringIO()
    writer = csv.DictWriter(out_buf, fieldnames=REQUIRED_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    out_buf.seek(0)

    if fixed_copies:
        logger.info(f"Corrected {fixed_copies} row(s) with copies=0 or empty.")
    if fixed_filenames:
        logger.info(f"Added missing .stl extension to {fixed_filenames} filename(s).")

    logger.info("meta.csv validated (in-memory).")
    return out_buf

# ------------------------------------------------------------
# ZIP packaging (RAM meta)
# ------------------------------------------------------------
import time
import zipfile
import os

def zip_with_limit(file_list, base_dir, zip_base_name, meta_buf, workdir, logger):
    """
    Create one or more ZIPs (900 MB each) for the given STL files.
    Files are read from `base_dir`, but ZIPs are always written to `workdir`.
    """
    import time
    MAX_ZIP_SIZE_MB = 900
    MAX_ZIP_SIZE_BYTES = MAX_ZIP_SIZE_MB * 1024 * 1024

    os.makedirs(workdir, exist_ok=True)
    zip_paths = []

    current_zip_index = 1
    current_zip = None
    current_size = 0
    start_time = time.time()

    def new_zip():
        nonlocal current_zip, current_zip_index, current_size
        zip_name = f"{zip_base_name}_part{current_zip_index}.zip"
        zip_path = os.path.join(workdir, zip_name)
        current_zip = zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_STORED)
        zip_paths.append(zip_path)
        current_size = 0
        logger.info(f"Started new archive: {zip_name}")
        print(f"Started new archive: {zip_name}")
        return zip_path

    # start the first archive
    new_zip()
    total_files = len(file_list)

    for i, filename in enumerate(file_list, start=1):
        full_path = os.path.join(base_dir, filename)
        if not os.path.isfile(full_path):
            continue
        rel_path = os.path.relpath(full_path, base_dir)
        file_size = os.path.getsize(full_path)

        if current_size + file_size > MAX_ZIP_SIZE_BYTES:
            # close and start new zip
            current_zip.close()
            elapsed = time.time() - start_time
            size_mb = current_size / (1024 * 1024)
            logger.info(f"Closed archive {current_zip_index} ({size_mb:.1f} MB in {elapsed:.1f}s, {size_mb/elapsed:.1f} MB/s)")
            current_zip_index += 1
            new_zip()
            start_time = time.time()

        with open(full_path, "rb") as f:
            data = f.read()
        current_zip.writestr(rel_path, data)
        current_size += file_size

        if i % 50 == 0 or i == total_files:
            print(f"Added {i}/{total_files} files so far...")
            logger.info(f"Added {i}/{total_files} files so far...")

    # finalize last archive
    if meta_buf:
        current_zip.writestr("meta.csv", meta_buf.getvalue())
    current_zip.close()

    elapsed = time.time() - start_time
    size_mb = current_size / (1024 * 1024)
    logger.info(f"Closed archive: {zip_paths[-1]} ({size_mb:.1f} MB in {elapsed:.1f}s, {size_mb/elapsed:.1f} MB/s)")
    print(f"Closed archive: {os.path.basename(zip_paths[-1])} ({size_mb:.1f} MB in {elapsed:.1f}s, {size_mb/elapsed:.1f} MB/s)")

    return zip_paths

# ------------------------------------------------------------
# Main
# ------------------------------------------------------------
def main(external_logger: logging.Logger | None = None) -> int:
    log_path = None
    workdir = get_working_dir()

    if external_logger:
        logger = setup_logger(None, external_logger)
    else:
        log_path = os.path.join(workdir, "converter_log.txt")
        logger = setup_logger(log_path)

    logger.info(f"Working directory: {workdir}")

    try:
        check_cancel()
        excel_path = next((os.path.join(workdir, f) for f in sorted(os.listdir(workdir))
                           if f.lower().endswith((".xlsx", ".xlsm"))), None)
        if not excel_path:
            logger.error("No Excel file found.")
            return 1

        basename = os.path.splitext(os.path.basename(excel_path))[0]
        logger.info(f"Found Excel: {excel_path}")

        check_cancel()
        csv_buffer = convert_excel_to_csv_buffer(excel_path, logger)
        meta_buf = validate_and_fix_meta_buffer(csv_buffer, logger)

        check_cancel()
        logger.info("=== START SCANNING FOR STL FILES ===")
        stl_folders, root_stls = [], []
        for entry in sorted(os.listdir(workdir)):
            p = os.path.join(workdir, entry)
            if os.path.isdir(p):
                if any(f.lower().endswith(".stl") for f in os.listdir(p)):
                    stl_folders.append(p)
            elif entry.lower().endswith(".stl"):
                root_stls.append(entry)

        total = len(root_stls) + sum(
            1 for f in stl_folders for fn in os.listdir(f) if fn.lower().endswith(".stl")
        )
        if total == 0:
            logger.error("No STL files found.")
            return 1

        logger.info(f"Found {total} STL files in total.")
        print(f"Found {total} STL files.\n")

        # --- Use tempdir (RAM-backed on Windows) for transient I/O
        with tempfile.TemporaryDirectory() as tmp:
            logger.info("=== START PACKAGING ===")
            for stl_folder in stl_folders:
                check_cancel()
                folder_name = os.path.basename(stl_folder)
                print(f"Packaging folder '{folder_name}'...")
                files = [f for f in os.listdir(stl_folder) if f.lower().endswith(".stl")]
                zip_paths = zip_with_limit(files, stl_folder, f"{basename}_{folder_name}",
                                           meta_buf, workdir, logger)
                for zp in zip_paths:
                    logger.info(f"Created archive: {zp} ({os.path.getsize(zp)/1024/1024:.2f} MB)")

            if root_stls:
                check_cancel()
                print(f"Packaging root STL files ({len(root_stls)} parts)...")
                zip_paths = zip_with_limit(root_stls, workdir, f"{basename}_root",
                                           meta_buf, workdir, logger)
                for zp in zip_paths:
                    logger.info(f"Created archive: {zp} ({os.path.getsize(zp)/1024/1024:.2f} MB)")

        logger.info("=== START CLEANUP ===")
        logger.info("All temporary data automatically cleared from memory.")
        return 0

    except RuntimeError as e:
        logger.warning(str(e))
        print(str(e))
        return 1
    except Exception as e:
        logger.exception("Unexpected error.")
        print(f"Unexpected error: {e}")
        return 1
    finally:
        # Automatic cleanup — also remove log file & cleaned sheets
        try:
            for f in os.listdir(workdir):
                if f.lower().startswith("_cleaned_") and f.lower().endswith((".xlsx", ".xlsm")):
                    try:
                        os.remove(os.path.join(workdir, f))
                        logger.info(f"Deleted temp file: {f}")
                    except Exception:
                        logger.warning(f"Could not delete {f}: {e}")

            if log_path and os.path.exists(log_path):
                try:
                    os.remove(log_path)
                    logger.info(f"Deleted log file: {log_path}")
                except Exception as e:
                    logger.warning(f"Could not delete log file: {e}")

        except Exception as cleanup_error:
            logger.warning(f"Cleanup encountered error: {cleanup_error}")

        logger.info("Converter finished.")

if __name__ == "__main__":
    sys.exit(main())